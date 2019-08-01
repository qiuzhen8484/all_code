# coding: utf-8

from openpyxl import load_workbook
import openpyxl
from shutil import copy
import os
import scipy.io as scio
import cv2

img_path = r'D:\Release0702\data\Mode_16'
pd_path = r'I:\threemode_excel\Mode_16\Lines_Mode_16'
dst_path = r'I:/threemode_excel/Mode_16/Mode_16_allexcel/'

img_list = []
for root, dirs, files in os.walk(img_path):
    for file in files:
        if os.path.splitext(file)[0][-2:] == '_1':
            img_list.append(file)

for img in img_list:
    if os.path.exists(os.path.join(dst_path, img[:-6]+'.xlsx')):
        # os.remove(os.path.join(dst_path, img[:-6]+'.xlsx'))
        continue
    copy('I:/1.xlsx', dst_path)
    os.rename(os.path.join(dst_path, '1.xlsx'), os.path.join(dst_path, img[:-6]+'.xlsx'))
    mm = cv2.imread(os.path.join(img_path, img[:-6] + '.jpg'), -1)
    wid = mm.shape[1]
    wb = load_workbook(os.path.join(dst_path[:-1], img[:-6]+'.xlsx'))
    sheets = wb.get_sheet_names()  # 从名称获取sheet
    booksheet = wb.get_sheet_by_name(sheets[0])
    for i in range(wid):
        booksheet.cell(row=1, column=i+2, value=str(i))

    mat = scio.loadmat(os.path.join(img_path, img))
    Lens_front_x = mat['A_lf_x'][0]
    Lens_front_y = mat['A_lf_y'][0]
    Lens_back_x = mat['A_lb_x'][0]
    Lens_back_y = mat['A_lb_y'][0]

    Cortex_front_x = mat['A_cf_x'][0]
    Cortex_front_y = mat['A_cf_y'][0]
    Cortex_back_x = mat['A_cb_x'][0]
    Cortex_back_y = mat['A_cb_y'][0]

    Nucleus_front_x = mat['A_nf_x'][0]
    Nucleus_front_y = mat['A_nf_y'][0]
    Nucleus_back_x = mat['A_nb_x'][0]
    Nucleus_back_y = mat['A_nb_y'][0]
    for i in range(wid):
        booksheet.cell(row=2, column=i + 2, value=str(-1))
        booksheet.cell(row=3, column=i + 2, value=str(-1))
        booksheet.cell(row=4, column=i + 2, value=str(-1))
        booksheet.cell(row=5, column=i + 2, value=str(-1))
        booksheet.cell(row=6, column=i + 2, value=str(-1))
        booksheet.cell(row=7, column=i + 2, value=str(-1))
    for i in range(len(Lens_front_x)):
        booksheet.cell(row=2, column=int(round(Lens_front_x[i])+2), value=str(Lens_front_y[i]))
    for i in range(len(Cortex_front_x)):
        booksheet.cell(row=3, column=int(round(Cortex_front_x[i])+2), value=str(Cortex_front_y[i]))
    for i in range(len(Nucleus_front_x)):
        booksheet.cell(row=4, column=int(round(Nucleus_front_x[i])+2), value=str(Nucleus_front_y[i]))
    for i in range(len(Nucleus_back_x)):
        booksheet.cell(row=5, column=int(round(Nucleus_back_x[i])+2), value=str(Nucleus_back_y[i]))
    for i in range(len(Cortex_back_x)):
        booksheet.cell(row=6, column=int(round(Cortex_back_x[i])+2), value=str(Cortex_back_y[i]))
    for i in range(len(Lens_back_x)):
        booksheet.cell(row=7, column=int(round(Lens_back_x[i])+2), value=str(Lens_back_y[i]))
    txt = open(os.path.join(pd_path, img[:-6]+'.txt'), 'r')
    lines = txt.readlines()
    id = 8
    for line in lines:
        item = line.split(sep=',')
        for i in range(wid):
            if item[i+1] != '0':
                booksheet.cell(row=id, column=i + 2, value=item[i + 1])
            else:
                booksheet.cell(row=id, column=i + 2, value=str(-1))
        id += 1
    wb.save(os.path.join(dst_path, img[:-6]+'.xlsx'))

