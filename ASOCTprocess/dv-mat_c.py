import os
import glob
import numpy as np
from PIL import Image
from shutil import move
import csv
from openpyxl import load_workbook
import openpyxl

path = r'F:\ASOCT_Data\Sick\201812'
path_save = 'F:/1200LGS'
path_2 = 'G:/0523parse'
data_mode = 'LGC'
interval = 0
patient_list = []

# workbook = load_workbook('C:/Users/cvter/Documents/WeChat Files/qz9607/Files/ASOCT_PatientInfo_By_CVTE(3).xlsx')
# sheets = workbook.get_sheet_names()         #从名称获取sheet
# booksheet = workbook.get_sheet_by_name(sheets[3])
# for i in range(772):
#     label = booksheet.cell(row=i+2, column=15).value
#     if label == 'Yes':
#         ID = booksheet.cell(row=i+2, column=2).value
#         patient_list.append(ID)

for file in os.listdir(path):
    for k in range(len(file)):
        if file[k] == '_':
            filename = file[:k]
            break
    if filename != '2146' and filename != '2150':
        continue
    # 获取样例的3dv文件路径
    path_f = os.path.join(path, file)
    path_xpf_list = glob.glob(os.path.join(path_f, '*.3dv'))
    path_3dv = "".join(path_xpf_list)
    if len(path_3dv) > 0:
        path_xpf_l = path_3dv[:-3] + 'xpf'
        path_xpf = os.path.join(path, file, path_xpf_l)
        with open(path_xpf, "r", encoding='utf-8') as f:
            for line1 in f:
                if len(line1) > 0:
                    # print(len(line1))
                    if len(line1) > 8:
                        if line1[0:8] == 'AB-Scan=':
                            width = int(line1[8:-1])
                    if len(line1) > 8:
                        if line1[0:8] == 'BC-Scan=':
                            num = int(line1[8:-1])
                    if len(line1) > 6:
                        if line1[0:6] == 'Depth=':
                            Depth = int(line1[6:-1])
                    if len(line1) > 4:
                        if line1[0:4] == 'Eye=':
                            name2 = line1[4]
                    if len(line1) > 10:
                        if line1[0:10] == 'PatientID=':
                            name1 = (line1[10:-1])
                    if len(line1) > 5:
                        if line1[0:5] == 'Date=':
                            name3 = line1[5:9] + line1[10:12] + line1[13:15]
                    if len(line1) > 5:
                        if line1[0:5] == 'Time=':
                            name4 = line1[5:7] + line1[8:10] + line1[11:13]
                    if len(line1) > 23:
                        if line1[0:13] == 'ScanTypeName=':
                            if data_mode == 'LRS':
                                if line1[0:24] == 'ScanTypeName=Lens Repeat':
                                    name5 = 'LRS'
                                else:
                                    name5 = 'XXX'
                            if data_mode == 'LBO':
                                if line1[0:26] == 'ScanTypeName=Lens Biometry':
                                    name5 = 'LBO'
                                else:
                                    name5 = 'XXX'
                            if data_mode == 'LGC':
                                if line1[0:24] == 'ScanTypeName=Lens Global':
                                    name5 = 'LGC'
                                else:
                                    name5 = 'XXX'

    if name5 == 'XXX':
        continue
    # if name1[0] == 'N' or name1[0] == 'n':
    #     if os.path.exists(os.path.join(path_2, name1)):
    #         pass
    #     else:
    #         os.makedirs(os.path.join(path_2, name1, 'left'))
    #         os.makedirs(os.path.join(path_2, name1, 'right'))
    # else:
    #     continue
    if os.path.exists(os.path.join(path_2, name1)):
        pass
    else:
        os.makedirs(os.path.join(path_2, name1, 'left'))
        os.makedirs(os.path.join(path_2, name1, 'right'))
    if name2 == 'L':
        path_save = os.path.join(path_2, name1) + '/left'
    else:
        path_save = os.path.join(path_2, name1) + '/right'

    namex = name1 + '_' + name3 + '_' + name4 + '_'+ name2 + '_CASIA2_' + name5 + '_'
    m = np.zeros((Depth * width), dtype=np.uint16)
    mm = np.zeros((Depth, width), dtype=np.uint16)
    if os.path.exists(path_3dv):
        with open(path_3dv, 'rb') as fid:
            data_array = np.fromfile(fid, np.uint16)
            for j in range(num):
                m = data_array[(j) * Depth * width: (j + 1) * Depth * width]
                for ii in range(Depth):
                    mm[ii, 0: width] = m[(ii) * width: (ii + 1) * width]

                # mm2 = mm
                # a1 = 65535 / 256 * 150
                # a2 = 65535 / 256 * 87
                # mm2 = np.where(mm2 > a2, 255 * (mm2 - a2) / (a1 - a2), 0)
                # mm2[mm2 > 255] = 255
                # mm2 = mm2.astype(np.uint8)
                k = j
                if interval == 0:
                    im = Image.fromarray(mm)
                    im = im.convert(mode='I')
                    j = "%03d" % j
                    name_save = namex + str(j) + '.png'
                    path_final = os.path.join(path_save, name_save)
                    if os.path.exists((path_final[:-4] + ".jpg")):
                        continue
                    im.save(path_final)
                    os.rename(path_final, path_final[:-4] + ".jpg")
                else:
                    im = Image.fromarray(mm)
                    im = im.convert(mode='I')
                    j = "%03d" % j
                    name_save = namex + str(j) + '.png'
                    path_final = os.path.join(path_save, name_save)
                    if k % interval == 0:
                        if os.path.exists((path_final[:-4] + ".jpg")):
                            continue
                        im.save(path_final)
                        os.rename(path_final, path_final[:-4] + ".jpg")

