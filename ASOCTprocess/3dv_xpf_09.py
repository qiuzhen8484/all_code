import os
import glob
import numpy as np
from PIL import Image
from shutil import move
import csv
from openpyxl import load_workbook
import openpyxl

path = 'G:/ASOCT_Data/Normal/201901'
path_save = 'H:/all_data_LRS'
path_2 = 'F:/Slit Lamp Image'
data_mode = 'LRS'
interval = 8
image_list = []
patient_list = []
flag = 'L'
finish = 0
ID_list = []
ID_list1 = []
txtlist = []

workbook = load_workbook('C:/Users/cvter/Documents/WeChat Files/qz9607/Files/ASOCT_PatientInfo_By_CVTE(3).xlsx')
sheets = workbook.get_sheet_names()         #从名称获取sheet
booksheet = workbook.get_sheet_by_name(sheets[3])
for i in range(772):
    label = booksheet.cell(row=i+2, column=15).value
    if label == 'Yes':
        ID = booksheet.cell(row=i+2, column=2).value
        patient_list.append(ID)

# for root, dirs, files in os.walk('F:/LGS新增数据/ASOCT'):
#     for file in files:
#         if os.path.splitext(file)[1] == '.jpg':
#             for j in range(len(file)):
#                 if file[j] == '_':
#                     ID = file[:j]
#                     break
#             if ID in patient_list:
#                 patient_list.remove(ID)
#                 continue
#             if ID[0] == 'n':
#                 if ('N' + ID[1:]) in patient_list:
#                     patient_list.remove(('N' + ID[1:]))
#
# for i in range(60):
#     image_list.append(patient_list[4*i])

# 计算例数
# for root, dirs, files in os.walk('F:/LRS新增数据/ASOCT1月'):
#     for file in files:
#         if os.path.splitext(file)[1] == '.jpg':
#             for j in range(len(file)):
#                 if file[j] == '_':
#                     ID = file[:j]
#                     break
#             if ID not in patient_list:
#                 patient_list.append(ID)
#
# f = open('C:/Users/cvter/Desktop/ASOCT测试集名单/LRSASOCT1月patentID.txt', 'x')
# for id in patient_list:
#     f.write(id + '\n')
# print(len(patient_list))

# for root, dirs, files in os.walk('D:/ASOCTpatientcount'):
#     for file in files:
#         txtlist.append(file)
#
# for i in range(len(txtlist)):
#     f = open(os.path.join('D:/ASOCTpatientcount', txtlist[i]), 'r')
#     lines = f.readlines()
#     for line in lines:
#         a = line.split()
#         if a[0] != 'patient_ID':
#             patient_list.append(a[0])
#
# for root, dirs, files in os.walk('F:/segfortest/good'):
#     for file in files:
#         if os.path.splitext(file)[1] == '.jpg':
#             for j in range(len(file)):
#                 if file[j] == '_':
#                     ID = file[:j]
#                     break
#             if ID not in patient_list:
#                 move(os.path.join('F:/segfortest/good', file), 'F:/segfortest/bad/')
#                 move(os.path.join('F:/LGS新增数据/good/label', (file[:-4] + '_1.mat')), 'F:/LGS新增数据/ASOCT1月/')


# 挑选部分代码
path1 = 'H:/all_data_LRS/S1'
for root, dirs, files in os.walk(path1):
    for file in files:
        if os.path.splitext(file)[1] == '.jpg':
            image_list.append(file)

for i in range(len(image_list)):
    for j in range(len(image_list[i])):
        if image_list[i][j] == '_':
            ID = image_list[i][:j]
            break
    if ID not in patient_list:
        if image_list[i+3][-6:-4] != '24':
            continue
        for k in range(4):
            move(os.path.join(path1, image_list[i+k]), 'H:/all_data_LRS/S/')
        flag = image_list[i][-20]
        patient_list = [ID]
        finish = 1
    if image_list[i][-20] != flag and (finish == 1):
        if image_list[i+3][-6:-4] != '24':
            continue
        for k in range(4):
            move(os.path.join(path1, image_list[i+k]), 'H:/all_data_LRS/S/')
        finish = 0

for file in os.listdir(path):
    # for k in range(len(file)):
    #     if file[k] == '_':
    #         filename = file[:k]
    #         break
    # 获取样例的3dv文件路径
    path_f = os.path.join(path, file)
    path_xpf_list = glob.glob(os.path.join(path_f, '*.3dv'))
    path_3dv = "".join(path_xpf_list)
    if len(path_3dv) > 0:
        path_xpf_l = path_3dv[:-3] + 'xpf'
        path_xpf = os.path.join(path, file, path_xpf_l)
        with open(path_xpf, "r", encoding='utf-8') as f:
            for line1 in f:
                if len(line1) > 0:
                    # print(len(line1))
                    if len(line1) > 8:
                        if line1[0:8] == 'AB-Scan=':
                            width = int(line1[8:-1])
                    if len(line1) > 8:
                        if line1[0:8] == 'BC-Scan=':
                            num = int(line1[8:-1])
                    if len(line1) > 6:
                        if line1[0:6] == 'Depth=':
                            Depth = int(line1[6:-1])
                    if len(line1) > 4:
                        if line1[0:4] == 'Eye=':
                            name2 = line1[4]
                    if len(line1) > 10:
                        if line1[0:10] == 'PatientID=':
                            name1 = (line1[10:-1])
                    if len(line1) > 5:
                        if line1[0:5] == 'Date=':
                            name3 = line1[5:9] + line1[10:12] + line1[13:15]
                    if len(line1) > 5:
                        if line1[0:5] == 'Time=':
                            name4 = line1[5:7] + line1[8:10] + line1[11:13]
                    if len(line1) > 23:
                        if line1[0:13] == 'ScanTypeName=':
                            if data_mode == 'LRS':
                                if line1[0:24] == 'ScanTypeName=Lens Repeat':
                                    name5 = 'LRS'
                                else:
                                    name5 = 'XXX'
                            if data_mode == 'LBO':
                                if line1[0:26] == 'ScanTypeName=Lens Biometry':
                                    name5 = 'LBO'
                                else:
                                    name5 = 'XXX'
                            if data_mode == 'LGC':
                                if line1[0:24] == 'ScanTypeName=Lens Global':
                                    name5 = 'LGC'
                                else:
                                    name5 = 'XXX'

    if name5 == 'XXX':
        continue
    if name1[0] == 's' or name1[0] == 'S':
        path_2 = path_save + '/S'
    elif name1[0] == 'M' or name1[0] == 'm':
        path_2 = path_save + '/M'
    elif name1[0] == 'c' or name1[0] == 'C':
        path_2 = path_save + '/C'
    else:
        path_2 = path_save + '/N'
    # if name1 in patient_list:
    #     continue
    # if name1[0] == 'n':
    #     ID = 'N' + name1[1:]
    # else:
    #     ID = name1
    # if ID not in image_list:
    #     continue

    namex = name1 + '_' + name3 + '_' + name4 + '_'+ name2 + '_CASIA2_' + name5 + '_'
    m = np.zeros((Depth * width), dtype=np.uint16)
    mm = np.zeros((Depth, width), dtype=np.uint16)
    if os.path.exists(path_3dv):
        with open(path_3dv, 'rb') as fid:
            data_array = np.fromfile(fid, np.uint16)
            for j in range(num):
                m = data_array[(j) * Depth * width: (j + 1) * Depth * width]
                for ii in range(Depth):
                    mm[ii, 0: width] = m[(ii) * width: (ii + 1) * width]

                # mm2 = mm
                # a1 = 65535 / 256 * 150
                # a2 = 65535 / 256 * 87
                # mm2 = np.where(mm2 > a2, 255 * (mm2 - a2) / (a1 - a2), 0)
                # mm2[mm2 > 255] = 255
                # mm2 = mm2.astype(np.uint8)
                k = j
                if interval == 0:
                    im = Image.fromarray(mm)
                    im = im.convert(mode='I')
                    j = "%03d" % j
                    name_save = namex + str(j) + '.png'
                    path_final = os.path.join(path_save, name_save)
                    if os.path.exists((path_final[:-4] + ".jpg")):
                        continue
                    im.save(path_final)
                    os.rename(path_final, path_final[:-4] + ".jpg")
                else:
                    im = Image.fromarray(mm)
                    im = im.convert(mode='I')
                    j = "%03d" % j
                    name_save = namex + str(j) + '.png'
                    path_final = os.path.join(path_2, name_save)
                    if k % interval == 0:
                        if os.path.exists((path_final[:-4] + ".jpg")):
                            continue
                        im.save(path_final)
                        os.rename(path_final, path_final[:-4] + ".jpg")
